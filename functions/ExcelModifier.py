import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename, column_to_modify, column_data, chart_boolean, name):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    column_name = sheet.cell(1, column_data)
    column_name.value = name
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, column_to_modify)
        corrected_price = cell.value * 0.5
        corrected_price_cell = sheet.cell(row, column_data)
        corrected_price_cell.value = corrected_price

    if chart_boolean:
        values = Reference(sheet, min_col=column_data, max_col=column_data, max_row=sheet.max_row, min_row=2)
        chart = BarChart()
        chart.add_data(values)
        sheet.add_chart(chart, 'e2')

    wb.save(filename)
